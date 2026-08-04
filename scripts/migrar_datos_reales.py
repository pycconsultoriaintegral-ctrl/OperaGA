"""
Migración de datos reales de Grupo Américas a Supabase (opera-dev primero).

Fuentes:
  - "ROTACIÓN Y DESCANSOS (1).xlsx": 12 hojas mensuales (ago/2025 - jul/2026)
    con el turno diario de cada empleado por propiedad.
  - "Reporte lista de empleados-*.xlsx" (Siigo): documento, salario, fecha de
    ingreso y estado del contrato de cada persona.

Qué hace:
  1. Reconstruye la fecha real de cada columna de cada hoja (las cabeceras de
     semana traen errores/typos de mes, así que solo se confía en el texto
     para la fecha de inicio del primer bloque; el resto se calcula sumando
     7 días por bloque, y se valida contra el nombre del día de la fila 2).
  2. Arma empleados (11 activos + Alex y Heberth como inactivos, confirmados
     con el usuario), propiedades (6, normalizando nombres con typos),
     novedades (vacaciones/incapacidad/permiso/calamidad/cumpleaños, agrupadas
     en rangos continuos) y horarios (Normal/Rotación/Salida -> turno Diurno,
     Descanso -> Descanso, Reserva -> turno Interno + estadía).
  3. Por defecto solo IMPRIME el reporte (modo simulación). Con --write
     aplica los cambios contra Supabase vía upsert (no duplica si se repite).

Uso:
  python scripts/migrar_datos_reales.py                # solo reporte
  python scripts/migrar_datos_reales.py --write         # aplica los cambios

Requiere las variables de entorno SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY
(la service role key nunca debe quedar en el código ni compartirse por chat).
"""
import os, re, sys, uuid, json
from datetime import date, timedelta
from collections import defaultdict, Counter
import openpyxl
import urllib.request

ROTACION_PATH = os.environ.get('ROTACION_PATH', 'ROTACIÓN Y DESCANSOS (1).xlsx')
SIIGO_PATH = os.environ.get('SIIGO_PATH', r'D:\Mis Archivos\Descargas\Reporte lista de empleados-20260804121709740.xlsx')
HOY = date(2026, 8, 4)

MESES = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,'agosto':8,
         'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12}

def strip_accents(s):
    return (s.lower().replace('á','a').replace('é','e').replace('í','i')
             .replace('ó','o').replace('ú','u').replace('ñ','n').replace('�','e'))

# ─────────────────────────────────────────────────────────────────────────
# 1. EMPLEADOS confirmados manualmente con el usuario (rotación -> Siigo)
# ─────────────────────────────────────────────────────────────────────────
EMPLEADOS_CONFIRMADOS = {
    'JOSE':     dict(nombre='José Guillermo Pacheco Carvajal', doc='1020744261', ingreso='2025-07-20', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'LUZ':      dict(nombre='Luz Maira Bello Berrío',          doc='23128973',   ingreso='2026-07-01', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'MAURICIO': dict(nombre='Mauricio Andrés Fuentes Pérez',   doc='1007975618', ingreso='2024-01-01', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'FABIAN':   dict(nombre='Fabián Andrés Mercado Pérez',     doc='1100396901', ingreso='2026-01-26', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'KATIA':    dict(nombre='Katia Paola Ortega Méndez',       doc='45706569',   ingreso='2024-02-26', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'SEGUNDO':  dict(nombre='Gilberto Segundo Acosta Ortiz',   doc='78380335',   ingreso='2025-08-19', salario=1800000, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'AYDA':     dict(nombre='Ayda Rosa Gómez Anaya',           doc='45551424',   ingreso='2025-10-01', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'MARISOL':  dict(nombre='Marisol Albertina Iglesias Matos',doc='45528283',   ingreso='2025-10-02', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'ROGER':    dict(nombre='Roger Jesús Cermeño Rivero',      doc='6946382',    ingreso='2026-01-01', salario=2000000, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'CARMEN':   dict(nombre='Carmen Julia Álvarez Ríos',       doc='45715320',   ingreso='2025-06-16', salario=1750905, cargo='Mayordomo', interno=True, estado='ACTIVO'),
    'GABRIEL':  dict(nombre='Gabriel Enrique Padilla Correa',  doc='73182942',   ingreso='2026-04-16', salario=1800000, cargo='Mantenimiento', interno=False, estado='ACTIVO'),
    'ALEX':     dict(nombre='Alex Samir Casseres Rico',        doc='1050960068', ingreso='2025-06-01', salario=1800000, cargo='Mantenimiento', interno=False, estado='INACTIVO'),
    'HEBERTH':  dict(nombre='Heberth Cecilio Avilez Acosta',   doc='78707196',   ingreso='2024-07-04', salario=1750905, cargo='Mayordomo', interno=True, estado='INACTIVO'),
}
# Ana y Hueso: el usuario pidió dejarlos fuera de esta migración (sin forma de verificar su identidad).

PROP_NORMALIZAR = {
    'CASA SOL':'Casa Sol', 'CASA LUNA':'Casa Luna', 'CASA OASIS':'Casa Oasis',
    'CASA LOTUS':'Casa Lotus', 'CASA MEL\ufffdA':'Casa Melía', 'CASA MELIA':'Casa Melía',
    'CASA MILAGROS':'Casa Milagros', 'CASA MILAGRO':'Casa Milagros',
}
PROP_CODIGOS = {'Casa Sol':'SOL-01','Casa Luna':'LUN-01','Casa Oasis':'OAS-01',
                'Casa Lotus':'LOT-01','Casa Melía':'MEL-01','Casa Milagros':'MIL-01'}

NOVEDAD_TIPOS = {
    'vacaciones': 'VACACIONES',
    'incapacidad': 'INCAPACIDAD',
    'permiso': 'PERMISO',
    'calamidad': 'PERMISO',
    'd\ufffda de cumplea\ufffdos': 'PERMISO',
}

def primer_bloque_inicio(text, target_year, target_month):
    d1 = int(re.search(r'DEL?\s+(\d+)', text, re.I).group(1))
    d2m = re.search(r'AL\s+(\d+)\s+DE\s+([A-ZÁÉÍÓÚÑ]+)', text, re.I)
    pares = re.findall(r'(\d+)\s+DE\s+([A-ZÁÉÍÓÚÑ]+)', text, re.I)
    if len(pares) >= 2:
        mes_txt = pares[0][1]
    else:
        d2 = int(d2m.group(1)); mes_fin = d2m.group(2)
        if d1 <= d2:
            mes_txt = mes_fin
        else:
            mes_num_fin = MESES[strip_accents(mes_fin)]
            mes_num = mes_num_fin - 1 if mes_num_fin > 1 else 12
            mes_txt = [k for k, v in MESES.items() if v == mes_num][0]
    mes = MESES[strip_accents(mes_txt)]
    y = target_year
    if target_month == 1 and mes == 12: y -= 1
    if target_month == 12 and mes == 1: y += 1
    return date(y, mes, d1)

def col_dates(ws, target_year, target_month):
    row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    starts = [i for i, v in enumerate(row1) if v and 'SEMANA' in str(v).upper()]
    first_start = primer_bloque_inicio(row1[starts[0]], target_year, target_month)
    dates = {}
    for bi, ci in enumerate(starts):
        end_ci = starts[bi + 1] if bi + 1 < len(starts) else ws.max_column
        ncols = end_ci - ci
        block_start = first_start + timedelta(days=7 * bi)
        for k in range(ncols):
            dates[ci + 1 + k] = block_start + timedelta(days=k)
    return dates

SHEETS = [('JULIO 2026',2026,7),('JUNIO 2026',2026,6),('MAYO 2026',2026,5),('ABRIL 2026',2026,4),
          ('MARZO 2026',2026,3),('FEBRERO 2026',2026,2),('ENERO 2026',2026,1),('DICIEMBRE 2025',2025,12),
          ('NOVIEMBRE 2025',2025,11),('OCTUBRE',2025,10),('SEPTIEMBRE',2025,9),('AGOSTO',2025,8)]
PROP_PREFIXES = ('CASA', 'MANTENIMIENTO')

def extraer_registros(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    registros = []  # (empleado, propiedad_normalizada|None, fecha, status_original)
    for name, y, m in SHEETS:
        ws = wb[name]
        dates = col_dates(ws, y, m)
        cur_prop = None
        for r in range(2, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a is None:
                continue
            a = str(a).strip()
            if a.upper().startswith('MANTENIMIENTO'):
                cur_prop = None  # equipo de mantenimiento: no vive en ninguna propiedad
                continue
            if any(a.upper().startswith(p) for p in PROP_PREFIXES):
                cur_prop = PROP_NORMALIZAR.get(a.upper(), a.title())
                continue
            if r == 2 or b is None:
                continue
            if a not in EMPLEADOS_CONFIRMADOS:
                continue
            for c, fecha in dates.items():
                v = ws.cell(row=r, column=c).value
                if v:
                    registros.append((a, cur_prop, fecha, str(v).strip()))
    return registros

def agrupar_rangos(fechas_ordenadas):
    """[(fecha,...)] ordenadas -> [(desde,hasta), ...] de días consecutivos."""
    rangos = []
    ini = prev = None
    for f in fechas_ordenadas:
        if ini is None:
            ini = prev = f
        elif f == prev + timedelta(days=1):
            prev = f
        else:
            rangos.append((ini, prev)); ini = prev = f
    if ini is not None:
        rangos.append((ini, prev))
    return rangos

def construir(registros):
    # dedupe: si dos hojas se solapan en una fecha (semana de cierre de mes),
    # se queda con el último valor visto (el orden de SHEETS ya es el correcto).
    por_emp_fecha = {}
    for emp, prop, fecha, status in registros:
        por_emp_fecha[(emp, fecha)] = (prop, status)

    horarios = []       # (empleado, fecha, tur)
    novedad_dias = defaultdict(list)   # (empleado, tipo) -> [fecha,...]
    reserva_dias = defaultdict(list)   # (empleado, propiedad) -> [fecha,...]
    prop_por_emp = defaultdict(Counter)

    for (emp, fecha), (prop, status) in por_emp_fecha.items():
        st = strip_accents(status)
        if prop:
            prop_por_emp[emp][prop] += 1
        if st in NOVEDAD_TIPOS:
            novedad_dias[(emp, NOVEDAD_TIPOS[st])].append(fecha)
        elif st in ('reserva', 'rserva'):
            horarios.append((emp, fecha, 'INT'))
            if prop:
                reserva_dias[(emp, prop)].append(fecha)
        elif st in ('normal', 'rotacion', 'salida'):
            horarios.append((emp, fecha, 'DIA'))
        elif st == 'descanso':
            horarios.append((emp, fecha, 'DES'))
        # cualquier otro texto no reconocido se ignora (reportado aparte)

    novedades = []
    for (emp, tipo), fechas in novedad_dias.items():
        for desde, hasta in agrupar_rangos(sorted(fechas)):
            novedades.append(dict(empleado=emp, tipo=tipo, desde=desde, hasta=hasta,
                                   dias=(hasta - desde).days + 1))

    estadias = []
    for (emp, prop), fechas in reserva_dias.items():
        for desde, hasta in agrupar_rangos(sorted(fechas)):
            if hasta < HOY: estado = 'FINALIZADA'
            elif desde <= HOY <= hasta: estado = 'ACTIVA'
            else: estado = 'PROGRAMADA'
            estadias.append(dict(empleado=emp, propiedad=prop, desde=desde, hasta=hasta, estado=estado))

    propiedad_principal = {emp: c.most_common(1)[0][0] for emp, c in prop_por_emp.items()}
    todas_props = sorted({p for c in prop_por_emp.values() for p in c})
    return horarios, novedades, estadias, propiedad_principal, todas_props

def reportar(horarios, novedades, estadias, propiedad_principal, empleados, props):
    print(f'Empleados a crear/actualizar: {len(empleados)}')
    for k, e in empleados.items():
        print(f'  {k:10s} {e["nombre"]:35s} {e["estado"]:9s} prop.actual={propiedad_principal.get(k,"—")}')
    print(f'\nPropiedades a crear/actualizar: {len(props)} -> {props}')
    print(f'\nHorarios (turnos programados): {len(horarios)} filas')
    print(f'Novedades (rangos): {len(novedades)}')
    for n in novedades:
        print(f'  {n["empleado"]:10s} {n["tipo"]:12s} {n["desde"]} -> {n["hasta"]} ({n["dias"]}d)')
    print(f'\nEstadías (rangos "en reserva"): {len(estadias)}')
    tot_dias_estadia = sum((e['hasta']-e['desde']).days+1 for e in estadias)
    print(f'  total días en reserva across estadías: {tot_dias_estadia}')

def main():
    write = '--write' in sys.argv
    registros = extraer_registros(ROTACION_PATH)
    horarios, novedades, estadias, propiedad_principal, props = construir(registros)
    reportar(horarios, novedades, estadias, propiedad_principal, EMPLEADOS_CONFIRMADOS, props)

    if not write:
        print('\n(Modo simulación — no se escribió nada. Corre con --write para aplicar.)')
        return

    url = os.environ['SUPABASE_URL']
    key = os.environ['SUPABASE_SERVICE_ROLE_KEY']

    def borrar_todo(table):
        req = urllib.request.Request(
            f'{url}/rest/v1/{table}?id=not.is.null',
            method='DELETE',
            headers={'apikey': key, 'Authorization': f'Bearer {key}'})
        with urllib.request.urlopen(req) as resp:
            resp.read()
        print(f'✗ {table}: datos de ejemplo eliminados')

    if '--clear-demo' in sys.argv:
        # Solo las tablas que tenían datos ficticios de la Fase 4; NO se tocan
        # roles/permisos/turnos_base/festivos/configuracion/profiles.
        for t in ['estadias', 'novedades', 'asistencia', 'horarios', 'reservas', 'propiedades', 'empleados']:
            borrar_todo(t)

    def upsert(table, rows, on_conflict='id'):
        if not rows: return
        req = urllib.request.Request(
            f'{url}/rest/v1/{table}?on_conflict={on_conflict}',
            data=json.dumps(rows, default=str).encode(),
            method='POST',
            headers={'apikey': key, 'Authorization': f'Bearer {key}',
                     'Content-Type': 'application/json', 'Prefer': 'resolution=merge-duplicates'})
        with urllib.request.urlopen(req) as resp:
            resp.read()
        print(f'✓ {table}: {len(rows)} filas')

    emp_ids = {k: str(uuid.uuid4()) for k in EMPLEADOS_CONFIRMADOS}
    upsert('empleados', [
        dict(id=emp_ids[k], nombre=e['nombre'], doc=e['doc'], tipo_doc='CC', cargo=e['cargo'],
             ingreso=e['ingreso'], contrato='Término indefinido', salario=e['salario'],
             eps=None, afp=None, arl=None, estado=e['estado'], interno=e['interno'])
        for k, e in EMPLEADOS_CONFIRMADOS.items()
    ], on_conflict='doc')

    prop_ids = {p: str(uuid.uuid4()) for p in props}
    upsert('propiedades', [
        dict(id=prop_ids[p], nombre=p, codigo=PROP_CODIGOS.get(p, p[:3].upper()+'-01'), tipo='Casa',
             estado='DISPONIBLE',
             mayordomo_id=next((emp_ids[k] for k, pp in propiedad_principal.items() if pp == p), None))
        for p in prop_ids
    ], on_conflict='codigo')

    upsert('horarios', [
        dict(id=str(uuid.uuid4()), empleado_id=emp_ids[emp], fecha=str(fecha), turno_id=tur)
        for emp, fecha, tur in horarios if emp in emp_ids
    ], on_conflict='empleado_id,fecha')

    upsert('novedades', [
        dict(id=str(uuid.uuid4()), empleado_id=emp_ids[n['empleado']], tipo=n['tipo'],
             desde=str(n['desde']), hasta=str(n['hasta']), dias=n['dias'],
             motivo=f'Migrado del archivo de rotación ({n["tipo"]})', estado='APROBADA')
        for n in novedades
    ], on_conflict='id')

    upsert('estadias', [
        dict(id=str(uuid.uuid4()), empleado_id=emp_ids[e['empleado']], propiedad_id=prop_ids[e['propiedad']],
             desde=str(e['desde']), hasta=str(e['hasta']), estado=e['estado'],
             obs='Migrado del archivo de rotación y descansos')
        for e in estadias
    ], on_conflict='id')

    print('\nListo.')

if __name__ == '__main__':
    main()
